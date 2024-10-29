from random import random
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from datetime import time
from colorama import Fore
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import re
import requests
import json
import os
import pandas as pd
import openpyxl


def get_band_response(brand_id="0"):
    num = 1  # 用于统计请求次数
    while True:
        headers = {
            "user-agent": UserAgent().random  # 随机获取ua
        }
        url = "https://car.autohome.com.cn/AsLeftMenu/As_LeftListNew.ashx"
        params = {
            "typeId": "1 ",
            "brandId": brand_id,
            "fctId": "0 ",
            "seriesId": "0"
        }
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response
        else:
            if num >= 5:
                print("请求超过5次，退出程序")
                break
            else:
                print("请求失败，正在重新请求...")
                num += 1
                time.sleep(1)


def main():
    while True:
        band = input("请输入汽车品牌：").strip()
        response = get_band_response()
        band_pattern = f"<a href=([^>]*?)><i[^>]*?></i>{band}<em>"
        band_info = re.search(band_pattern, response.text)
        if not band_info:
            print("该品牌不存在，请重新输入")
            continue
        else:
            band_href = band_info.group(1)
            band_id = re.findall(r'/price/brand-(\d+).html', band_href)[0]
            print(F"{band} 品牌id为：", band_id)
            resp_brand = get_band_response(brand_id=band_id)
            # 上面得到了品牌页面的响应数据后，即可往下解析出该品牌下的各个车系的名称的id值
            parse_series(band, response=resp_brand)

            break


def parse_series(band, response):
    html = re.findall(r'document.writeln\("(.*)"\)', response.text)
    html = "".join(html)
    soup = BeautifulSoup(html, "html.parser")
    data_list = soup.select(".current > dl > dd > a")
    still_sell = [i for i in data_list if "停售" not in i.get_text(strip=True)]
    stop_sell = [i for i in data_list if "停售" in i.get_text(strip=True)]
    print(
        f"该品牌共找到{len(data_list)}个车型，其中，在售车型共{len(still_sell)}个，已停售车型共{len(stop_sell)}个车型（停售系列车型无配置信息）。")
    print(
        "----------------------------------------------\n在售车型列表如下：\n----------------------------------------------")
    series_dict = {}
    for still_index, still_data in enumerate(still_sell, start=1):
        series_name = still_data.contents[0].text.strip()
        href = still_data.get("href")
        series_id = re.findall(r'/price/series-(\d+).html', href)[0]
        series_dict[series_id] = series_name
        print(f"序号：{still_index}\t车型：{series_name}\t车型id：{series_id}")

    while True:
        choice = input(Fore.RED + "\n请输入需要下载的车型id：").strip()
        if choice in series_dict.keys():
            # 以下为获取配置的逻辑函数
            # 构建配置页url
            series_name = series_dict[choice]
            series_url = "https://car.autohome.com.cn/config/series/{}.html".format(choice)
            print(Fore.CYAN + f"---正在下载{band}-{series_name}，车型id为：{choice}，配置链接为：{series_url}")
            # 获取当前车系的响应数据，即配置，此时的配置信息是不完整的，其中的部分数据是隐藏的，需要解密
            response = get_response(choice)
            if "抱歉" in response.text and "暂无相关数据" in response.text:
                print(Fore.RED + "该系列车暂无配置信息")

            # 字典格式的配置信息
            resp_dict = json.loads(response.text)
            # print(resp_dict)
            # 获取多列配置数据
            all_info = get_car_config(resp_dict)
            df = pd.DataFrame(all_info)
            # print(df)
            # 根据要求，提取出车系的上市年份，构建文件名
            # year_re = re.search(r'(\d{4}款)', list(car_config["车型名称"].values())[0].replace(" ", ""))
            # year = year_re.group(1) if year_re else ""
            # excel_name = f"{band}_{series_name}_{year}.xlsx"
            excel_name = f"{band}_{series_name}.xlsx"
            # 保存到excel文件中
            save_to_excel(all_info, folder=band, filename=excel_name)

            break
        else:
            print("输入的车型id不存在，请重新输入。")
            continue
    input("请按任意键关闭程序...")


def get_response(series_id="0"):
    num = 1  # 用于统计请求次数
    while True:
        headers = {
            "user-agent": UserAgent().random  # 随机获取ua
        }
        url = "https://car-web-api.autohome.com.cn/car/param/getParamConf"
        params = {
            "mode": "1",
            "site": "1",
            "seriesid": series_id
        }
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response
        else:
            if num >= 5:
                print("请求超过5次，退出程序")
                break
            else:
                print("请求失败，正在重新请求...")
                num += 1
                time.sleep(1)


# 清洗数据
def get_car_config(config_dic):
    # 获取配置项列表
    allconfig = []
    # 初始化itemname列表
    configname_list = []
    # 遍历titlelist数组
    for title in config_dic['result']['titlelist']:
        # 遍历items数组
        for item in title['items']:
            # 提取itemname并添加到列表中
            configname_list.append(item['itemname'])

    allconfig.append(configname_list)

    # 获取配置数据
    for data in config_dic['result']['datalist']:
        configvalue_list = []
        # 注意多个数据调整格式，颜色数据另外处理
        for valueitem in data['paramconflist']:
            value_list = []
            if valueitem.get('itemname') != '':
                configvalue_list.append(valueitem['itemname'])
            elif not valueitem.get('sublist'):
                configvalue_list.append('-')
            else:
                stri = []
                for multivalue in valueitem['sublist']:
                    stri.append(multivalue['value'] + multivalue['name'])
                # 连成一个文本串，不要列表形式防止多余'[]'
                stro = '\n'.join(stri)
                configvalue_list.append(stro)
        allconfig.append(configvalue_list)
        # 颜色之后处理一下再
    return allconfig


# 保存数据
def save_to_excel(data, folder, filename):
    if not os.path.exists(folder):
        os.mkdir(folder)
    df = pd.DataFrame(data)
    # df.T是将表格的行和列进行倒置的操作
    excel_path = f"{folder}/{filename}"
    df.T.to_excel(excel_path, index=False, header=False)

    # 使用openpyxl打开Excel文件，修改单元格对齐方式以启用换行
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 遍历所有单元格，启用换行和垂直居中
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    # 设置列宽
    num_columns = df.shape[0]
    for col in range(1, num_columns + 1):
        sheet.column_dimensions[chr(64 + col)].width = 20

    # 保存对工作簿的更改
    workbook.save(excel_path)

    print(Fore.GREEN + "配置下载完成，保存到文件------> ", f"{folder}/{filename}")


if __name__ == '__main__':
    main()